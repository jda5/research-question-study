
# ------ LITERATURE FINDER ------

# This program will find papers published between 2016 and 2021 in (user-defined) journals hosted on Elsevier, JStor and Springer.

# The program will only search for peer-reviewed studies, ie. book reviews, citations, editorials, author corrections, etc. are omitted.

# The program will retrieve the web address for each paper in the journal and save them to excel spreadsheet. The spreadsheet is to be used with the accompanying file
# "paper_scanner.py"

# As most papers are divided into volumes, two functions are created for each website: one that navigates the journal's main page, and another than navigates the
# issues within each volume. These functions follow the naming convention {journal_name}_main and {journal_name}_volume.

# The program uses the Firefox browser to access the papers. This will need to be installed prior to its excecution.
# See https://www.mozilla.org/ for more details.

# DISCLAIMER: This program was used to access articles at a specific point in time. There is no guarantee that the program will work in the future as the websites
# may have altered their layout.

# --- Imports ---

# requests, bs4, openpyxl and selenium must be installed before running the program. These can be installed by simply entering pip install {module_name} in the command
# line. See https://pip.pypa.io/en/stable/ for more details.

import requests
from bs4 import BeautifulSoup
import openpyxl
from selenium import webdriver
from time import sleep
import re

# User agent header to tell the webpage that a legitimate browser is accessing the page. This has been classified, and should be changed by the user.
headers = {
    'User-Agent': '****** CHANGE_ME *******'
}


# --- Elsevier ---

def elsevier_main(url: str):
    browser = webdriver.Firefox()
    browser.get(url)
    sleep(6)
    browser.find_element_by_xpath('/html/body/div[4]/div/div/div/main/div[3]/div/section[2]/div/div/ol/li['
                                  '2]/button/span').click()
    sleep(1)
    browser.find_element_by_xpath('/html/body/div[4]/div/div/div/main/div[3]/div/section[2]/div/div/ol/li['
                                  '3]/button/span').click()
    sleep(1)
    browser.find_element_by_xpath('/html/body/div[4]/div/div/div/main/div[3]/div/section[2]/div/div/ol/li['
                                  '4]/button/span').click()
    sleep(1)
    browser.find_element_by_xpath('/html/body/div[4]/div/div/div/main/div[3]/div/section[2]/div/div/ol/li['
                                  '5]/button/span').click()
    sleep(1)
    soup = BeautifulSoup(browser.page_source, 'lxml')
    browser.close()
    links = []
    for volume in soup.find_all(class_="accordion-panel-content u-padding-s-ver js-accordion-panel-content"):
        for a in volume.find_all(class_="anchor js-issue-item-link text-m"):
            extension = a.get("href", None)
            links.append(f'https://www.sciencedirect.com/{extension}')

    return links


def elsevier_volume(url: str):
    global row
    global sheet
    global headers
    date_regex = re.compile(r'\w+ \d{4}')
    res = requests.get(url, headers=headers)
    try:
        res.raise_for_status()
    except Exception as exc:
        print(exc)
        print(f'\nInvalid URL: {url}')
    soup = BeautifulSoup(res.text, 'lxml')
    j_info = soup.find(class_="col-md-16 u-padding-l-right-from-md u-margin-l-bottom")
    volume = j_info.find(class_="u-text-light u-h1 js-vol-issue").text
    year = j_info.find(class_="js-issue-status text-s").text
    print(year)
    for article in soup.find_all(class_="js-article-list-item article-item u-padding-xs-top u-margin-l-bottom"):
        article_info = article.find(class_="js-article article-content").find(class_="article-info text-xs").text
        if 'Research article' in article_info:
            link = article.find(class_="js-article-link u-margin-xs-top u-margin-s-bottom article-item-link")
            extension = link.find(class_="anchor pdf-download u-margin-l-right text-s").get("href", None)
            sheet.cell(row=row, column=3, value=f'https://www.sciencedirect.com{extension}')
            sheet.cell(row=row, column=2, value=f'{volume}, {date_regex.search(year).group()}')
            row += 1

# --- JStor ---


def jstor_main(url: str):
    browser = webdriver.Firefox()
    browser.get(url)
    sleep(3)
    browser.find_element_by_xpath('//*[@id="content"]/div[1]/div[2]/div[3]/div/div/dl/dl[2]/dt').click()
    sleep(1)
    browser.find_element_by_xpath('//*[@id="content"]/div[1]/div[2]/div[3]/div/div/dl/dl[1]/dt').click()
    sleep(1)
    soup = BeautifulSoup(browser.page_source, 'lxml')
    browser.close()
    volumes = soup.find_all(class_="pbxl mll")
    addresses = []
    for volume in volumes:
        for content in volume.contents:
            try:
                link = content.a
            except:
                continue
            if "2016" in link.text or "2017" in link.text or "2018" in link.text or "2019" in link.text or "2020" in link.text:
                extension = link.get("href", None)
                addresses.append(f'https://www.jstor.org{extension}')
    return addresses


def jstor_volume(url: str):
    global sheet
    global row
    res = requests.get(url, headers=headers)
    try:
        res.raise_for_status()
    except Exception as exc:
        print(exc)
        print(f'\nInvalid URL: {url}')
        return
    soup = BeautifulSoup(res.text, 'lxml')
    journal_title = soup.find(class_="normal medium-heading").text
    journal_information = soup.find(class_="issue").text
    for section in soup.find_all(class_="small-12 columns mvs"):
        section_type = section.find(class_="mln").text
        if 'BOOK REVIEW' in section_type: # Keyword exlusions
            continue
        elif 'EDITORIAL' in section_type:
            continue
        elif 'Cite this Item' in section_type:
            continue
        elif 'INTRODUCTION' in section_type:
            continue
        elif '50th ANNIVERSARY REFLECTION' in section_type:
            continue
        elif 'RESEARCH COMMITTEE' in section_type:
            continue
        elif 'RESEARCH COMMENTARY' in section_type:
            continue
        else:
            for article in section.find_all(class_="media-body media-object-section main-section"):
                extensions = article.find(class_="tt-track small-heading inline").get("href", None)
                sheet.cell(row=row, column=4, value=f'https://www.jstor.org{extensions}')
                sheet.cell(row=row, column=3, value=journal_information)
                sheet.cell(row=row, column=2, value=journal_title)
                sheet.cell(row=row, column=1, value=section_type)
                row += 1

# --- Springer ---

def springer_main(url: str):
    global sheet
    global row
    res = requests.get(url, headers=headers)
    try:
        res.raise_for_status()
    except Exception as exc:
        print(exc)
        print(f'\nInvalid URL: {url}')
        return
    soup = BeautifulSoup(res.text, 'lxml')
    journal_title = soup.find(id="journalTitle").text
    journal_edition = soup.find(class_="app-volumes-and-issues__info").find(class_="u-mb-8").text
    for article in soup.find_all(class_="c-list-group__item"):
        body = article.find(class_="c-card__body app-card-body")
        meta_data = body.find(class_="c-meta u-text-xs")
        article_type = meta_data.find(class_="c-meta__item").text[14:]
        if article_type == "Announcement":  # Keyword exlusions
            continue
        elif article_type == "BookReview":
            continue
        elif article_type == "EditorialNotes":
            continue
        elif article_type == "Author Correction":
            continue
        elif article_type == "Acknowledgments":
            continue
        elif article_type == "Editorial":
            continue
        elif article_type == "Obituary":
            continue
        elif article_type == "Letter":
            continue
        elif article_type == "Commentary":
            continue
        elif article_type == "Commentary Paper":
            continue
        else:
            sheet.cell(row=row, column=4, value=article.find(itemprop="url").get("href"))
            sheet.cell(row=row, column=3, value=journal_edition)
            sheet.cell(row=row, column=2, value=journal_title)
            sheet.cell(row=row, column=1, value=article_type)
            row += 1


def springer_volume(url: str):
    res = requests.get(url, headers=headers)
    try:
        res.raise_for_status()
    except Exception as exc:
        print(exc)
        print(f'\nInvalid URL: {url}')
    soup = BeautifulSoup(res.text, 'lxml')
    links = []
    for volume in soup.find_all(class_="app-section"):
        for year in ('2016', '2017', '2018', '2019', '2020'):
            if year in volume.find(class_="app-section__heading app-section__heading--space-between-at-sm").text:
                for issue in volume.find_all(class_="c-list-group__item"):
                    path = issue.find(class_="u-interface-link u-text-sans-serif u-text-sm").get("href", None)
                    links.append(f'https://link.springer.com{path}')
    return links


if __name__ == '__main__':

    
    wb = openpyxl.Workbook() # Start a new workbook
    sheet = wb.active
    row = 1
    address = jstor_main('https://www.jstor.org/journal/jresematheduc') # --- Paste the journal URL here ---
    for locator in address:
        jstor_volume(locator) # Locates the papers in each volume

    wb.save('JStor Articles.xlsx') # Saves the excel file. Change the name of the file to prevent overlapping.

